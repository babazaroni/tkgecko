from unittest.util import sorted_list_difference

from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Side,Border
#from onedrive import  upload_and_link

#import globals as glb



class Cell:
    def __init__(self,value = None,border = None, font = None,fill = None):
        self.value = value
        self.border = border if border else Border()
        #self.font = font if font else Font()
        #self.border = border
        self.font = font
        self.fill = fill


class Canvas:
    def __init__(self):
        self.cmds = []
        self.max_row = 0
        self.max_col = 0
        self.view  = None

    def set(self,row,col,value = None,border = None, font = None,fill = None):
        cell = Cell(value=value,border=border,font=font,fill=fill)
        cmd = {'row':row,'col':col,'cell':cell}
        self.cmds.append(cmd)
        self.max_row = max(self.max_row,row);self.max_col = max(self.max_col,col)


    def allocate(self):

        self.view = [[ Cell() for col in range(self.max_col+1)] for row in range(self.max_row+1)]

        for cmd in self.cmds:
            self.view[cmd['row']][cmd['col']] = cmd['cell']

    def set_row(self,row,value = None,border = None, font = None,fill = None):
        new_cell = Cell(value=value,border=border,font=font,fill=fill)
        cell_row = self.view[row]
        for cell in cell_row:
            self.transfer_values(new_cell, cell)

    def set_col(self,col,value = None,border = None, font = None,fill = None):
        new_cell = Cell(value=value,border=border,font=font,fill=fill)
        for r in self.view:
            self.transfer_values(new_cell, r[col])

    def transfer_values(self, new_cell, cell):
        if new_cell.value:
            cell.value = new_cell.value
        if new_cell.border:
            if new_cell.border.top and not cell.border.top:
                cell.border.top = new_cell.border.top
            if new_cell.border.bottom and not cell.border.bottom:
                cell.border.bottom = new_cell.border.bottom
            if new_cell.border.left and not cell.border.left:
                cell.border.left = new_cell.border.left
            if new_cell.border.right and not cell.border.right:
                cell.border.right = new_cell.border.right

        if new_cell.fill:
            cell.fill = new_cell.fill

    def all(self,new_cell):
        for row in self.view:
            for cell in row:
                self.transfer_values(new_cell, cell)

    def border(self,side):
        for r,row in enumerate(self.view):
            for c,cell in enumerate(row):
                if r==0:
                    cell.border.top = side
                if r==self.max_row:
                    cell.border.bottom = side
                if c==0:
                    cell.border.left = side
                if c==self.max_col:
                    cell.border.right = side

    def render(self,sheet, start_row, start_col):
        for r,row in enumerate(self.view):
            for c,cell in enumerate(row):
                loc = '{}{}'.format(chr(ord('A')+start_col+c),start_row+r+1)

                sheet[loc] = cell.value
                sheet[loc].border = cell.border
                sheet[loc].font = cell.font
                if cell.fill:
                    sheet[loc].fill = cell.fill



def row_color(sheet,row_num,width_cell,color):
    sheet["{}{}".format(width_cell,row_num+1)] = "  " #put space so we set the width of the sheet
    row = sheet[row_num+1]
    for cell in row:
        cell.fill = PatternFill('solid',start_color=color)



